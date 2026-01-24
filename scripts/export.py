#!/usr/bin/env python3
"""
Export company data to CSV, XLSX, and JSON formats.
Reads markdown files from content/company/ and validates against taxonomies.
"""

import json
import sys
from pathlib import Path

import frontmatter
import pandas as pd
import yaml
from openpyxl.utils import get_column_letter

# Paths
ROOT = Path(__file__).parent.parent
CONTENT_DIR = ROOT / "content" / "company"
DATA_DIR = ROOT / "data"
STATIC_DIR = ROOT / "static"


def load_taxonomies() -> dict:
    """Load taxonomy definitions from data/taxonomies.yaml"""
    tax_file = DATA_DIR / "taxonomies.yaml"
    if not tax_file.exists():
        print(f"Warning: {tax_file} not found", file=sys.stderr)
        return {}

    with open(tax_file) as f:
        return yaml.safe_load(f) or {}


def load_companies() -> list[dict]:
    """Load all company markdown files."""
    companies = []

    for md_file in sorted(CONTENT_DIR.glob("*.md")):
        # Skip index files
        if md_file.name.startswith("_"):
            continue

        try:
            post = frontmatter.load(str(md_file))
            company = dict(post.metadata)
            company["_content"] = post.content
            company["_file"] = md_file.name
            # Derive slug from filename (Hugo convention)
            company["slug"] = md_file.stem
            companies.append(company)
        except Exception as e:
            print(f"Error loading {md_file.name}: {e}", file=sys.stderr)

    return companies


def validate_company(company: dict, taxonomies: dict) -> list[str]:
    """Validate a company against taxonomy values. Returns list of errors."""
    errors = []
    filename = company.get("_file", "unknown")

    # Required fields
    if not company.get("name"):
        errors.append(f"{filename}: missing required field 'name'")

    # Validate taxonomy values (frontmatter contains keys, not display names)
    # Taxonomy format: { short_key: "Full Name" }
    def get_valid_keys(tax_dict):
        """Extract valid keys from taxonomy dict."""
        if not tax_dict:
            return []
        return list(tax_dict.keys())

    taxonomy_fields = [
        ("stakeholders", get_valid_keys(taxonomies.get("stakeholders", {}))),
        (
            "capability_streams",
            get_valid_keys(taxonomies.get("capability_streams", {})),
        ),
        (
            "capability_domains",
            get_valid_keys(taxonomies.get("capability_domains", {})),
        ),
        (
            "industrial_capabilities",
            get_valid_keys(taxonomies.get("industrial_capabilities", {})),
        ),
        ("regions", get_valid_keys(taxonomies.get("regions", {}))),
    ]

    for field, valid_keys in taxonomy_fields:
        company_values = company.get(field, []) or []
        for val in company_values:
            if val not in valid_keys:
                errors.append(f"{filename}: invalid {field} key '{val}'")

    return errors


def companies_to_dataframe(companies: list[dict], taxonomies: dict) -> pd.DataFrame:
    """Convert companies to a pandas DataFrame.

    Converts taxonomy keys to display names for human-readable output.
    Column order optimized: identity -> location -> taxonomies -> flags -> content.
    """

    # Helper to convert keys to display names
    def keys_to_names(keys: list[str], tax_name: str) -> list[str]:
        mapping = taxonomies.get(tax_name, {})
        return [mapping.get(k, k) for k in keys]

    rows = []
    for c in companies:
        row = {
            # Identity (most important first)
            "name": c.get("name", ""),
            "website": c.get("website", ""),
            # Location
            "regions": "; ".join(keys_to_names(c.get("regions") or [], "regions")),
            "address": c.get("address", ""),
            # Capabilities (most searched)
            "capability_streams": "; ".join(
                keys_to_names(c.get("capability_streams") or [], "capability_streams")
            ),
            "capability_domains": "; ".join(
                keys_to_names(c.get("capability_domains") or [], "capability_domains")
            ),
            "industrial_capabilities": "; ".join(
                keys_to_names(
                    c.get("industrial_capabilities") or [], "industrial_capabilities"
                )
            ),
            # Classification
            "stakeholders": "; ".join(
                keys_to_names(c.get("stakeholders") or [], "stakeholders")
            ),
            # Flags
            "is_sme": c.get("is_sme", False),
            "is_prime": c.get("is_prime", False),
            "is_indigenous_owned": c.get("is_indigenous_owned", False),
            "is_veteran_owned": c.get("is_veteran_owned", False),
            # Contact
            "phone": c.get("phone", ""),
            "email": c.get("email", ""),
            # Geo
            "latitude": c.get("latitude"),
            "longitude": c.get("longitude"),
            # Full markdown content
            "markdown": (c.get("_content") or "").strip(),
            # Internal reference
            "slug": c.get("slug", ""),
        }
        rows.append(row)

    return pd.DataFrame(rows)


def companies_to_json(companies: list[dict], taxonomies: dict) -> list[dict]:
    """Convert companies to JSON-serializable format.

    Preserves taxonomy keys (not display names) for programmatic access.
    """
    result = []
    for c in companies:
        item = {
            "slug": c.get("slug", ""),
            "name": c.get("name", ""),
            "website": c.get("website", ""),
            "address": c.get("address", ""),
            "phone": c.get("phone", ""),
            "email": c.get("email", ""),
            "latitude": c.get("latitude"),
            "longitude": c.get("longitude"),
            # Full markdown content
            "markdown": (c.get("_content") or "").strip(),
            # Taxonomies as arrays (keys, not display names)
            "regions": c.get("regions") or [],
            "stakeholders": c.get("stakeholders") or [],
            "capability_streams": c.get("capability_streams") or [],
            "capability_domains": c.get("capability_domains") or [],
            "industrial_capabilities": c.get("industrial_capabilities") or [],
            # Flags
            "is_sme": c.get("is_sme", False),
            "is_prime": c.get("is_prime", False),
            "is_indigenous_owned": c.get("is_indigenous_owned", False),
            "is_veteran_owned": c.get("is_veteran_owned", False),
        }
        result.append(item)

    return result


def export_csv(df: pd.DataFrame, output_path: Path) -> None:
    """Export DataFrame to CSV."""
    df.to_csv(output_path, index=False)
    print(f"Exported {len(df)} companies to {output_path}")


def export_xlsx(df: pd.DataFrame, output_path: Path) -> None:
    """Export DataFrame to XLSX with formatting and hyperlinks."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Companies")

        worksheet = writer.sheets["Companies"]

        # Find website column index (1-based for openpyxl)
        website_col_idx = list(df.columns).index("website") + 1

        # Add hyperlinks to website column
        for row_idx, url in enumerate(
            df["website"], start=2
        ):  # Start at row 2 (after header)
            if url and isinstance(url, str) and url.startswith("http"):
                cell = worksheet.cell(row=row_idx, column=website_col_idx)
                cell.hyperlink = url
                cell.style = "Hyperlink"

        # Auto-adjust column widths
        for idx, col in enumerate(df.columns):
            col_letter = get_column_letter(idx + 1)
            # Handle NaN values properly
            col_data = df[col].fillna("").astype(str)
            max_length = max(col_data.map(len).max(), len(col))
            # Cap widths: 80 for markdown, 50 for others
            max_width = 80 if col == "markdown" else 50
            worksheet.column_dimensions[col_letter].width = min(
                max_length + 2, max_width
            )

        # Freeze header row
        worksheet.freeze_panes = "A2"

    print(f"Exported {len(df)} companies to {output_path}")


def export_json(companies: list[dict], output_path: Path) -> None:
    """Export companies to JSON."""
    with open(output_path, "w") as f:
        json.dump(companies, f, indent=2)
    print(f"Exported {len(companies)} companies to {output_path}")


def main():
    """Main export function."""
    print("Loading taxonomies...")
    taxonomies = load_taxonomies()

    print(f"Loading companies from {CONTENT_DIR}...")
    companies = load_companies()
    print(f"Found {len(companies)} companies")

    # Validate
    print("Validating companies...")
    all_errors = []
    for company in companies:
        errors = validate_company(company, taxonomies)
        all_errors.extend(errors)

    if all_errors:
        print(f"\nValidation errors ({len(all_errors)}):", file=sys.stderr)
        for error in all_errors[:20]:  # Show first 20
            print(f"  - {error}", file=sys.stderr)
        if len(all_errors) > 20:
            print(f"  ... and {len(all_errors) - 20} more", file=sys.stderr)
        print("\nExport will continue but data may be incomplete.", file=sys.stderr)

    # Convert to DataFrame
    df = companies_to_dataframe(companies, taxonomies)

    # Sort by name
    df = df.sort_values("name", ignore_index=True)

    # Export
    STATIC_DIR.mkdir(parents=True, exist_ok=True)

    export_csv(df, STATIC_DIR / "companies.csv")
    export_xlsx(df, STATIC_DIR / "companies.xlsx")

    # JSON export (uses original companies data, sorted)
    json_data = companies_to_json(
        sorted(companies, key=lambda c: c.get("name", "").lower()), taxonomies
    )
    export_json(json_data, STATIC_DIR / "companies.json")

    print("\nExport complete!")
    return 0  # Always succeed - validation is informational


if __name__ == "__main__":
    sys.exit(main())
