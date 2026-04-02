#!/usr/bin/env python3
"""
Preprocess company data for directory site.

Generates:
  - static/maps/*.png: Static minimap images for companies
  - static/maps/terms/*.png: Static map images for taxonomy terms
  - static/companies.csv, .xlsx, .json: Export files

Usage:
    uv run python scripts/preprocess.py
"""

import hashlib
import json
from typing import Any
from pathlib import Path

import httpx
import frontmatter
import pandas as pd
import markdown_it
from openpyxl.utils import get_column_letter
from mlnative import Map, from_latlng

from config import (
    cache,
    params,
    COMPANY_DIR,
    DATA_DIR,
    MAPS_DIR,
    STATIC_DIR,
    TERM_MAPS_DIR,
    TAXONOMIES,
    MAP_WIDTH,
    MAP_HEIGHT,
    MAP_MARKER_LAYERS,
    load_taxonomies,
    make_progress,
)

# Initialize markdown-it for HTML conversion
md = markdown_it.MarkdownIt()

# --- Map rendering ---


def fetch_map_style() -> dict:
    """Fetch map style JSON with cache and short timeout."""
    cache_key = f"map-style:{params['mapStyleUrl']}"
    if cache_key in cache:
        return cache[cache_key]

    response = httpx.get(params["mapStyleUrl"], follow_redirects=True, timeout=10.0)
    response.raise_for_status()
    style = response.json()
    cache[cache_key] = style
    return style


def setup_map(width: int = MAP_WIDTH, height: int = MAP_HEIGHT) -> tuple[Map, dict]:
    """Create a Map instance and load style with marker layers."""
    style = fetch_map_style()
    style["sources"]["markers"] = {
        "type": "geojson",
        "data": {"type": "FeatureCollection", "features": []},
    }
    style["layers"].extend(MAP_MARKER_LAYERS)
    return Map(width, height, pixel_ratio=2), style


def map_input_digest(
    locations: list[tuple[float, float]],
    *,
    max_zoom: float,
    zoom_out: float,
) -> str:
    """Return a stable digest for rendered map inputs."""
    payload = {
        "locations": sorted((round(lat, 6), round(lng, 6)) for lat, lng in locations),
        "max_zoom": max_zoom,
        "zoom_out": zoom_out,
        "style_url": params["mapStyleUrl"],
        "marker_layers": MAP_MARKER_LAYERS,
        "width": MAP_WIDTH,
        "height": MAP_HEIGHT,
        "pixel_ratio": 2,
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode()
    return hashlib.sha256(encoded).hexdigest()


def sidecar_path(output: Path) -> Path:
    """Return sidecar path storing the last rendered input digest."""
    return output.with_suffix(".sha256")


def needs_render(output: Path, digest: str) -> bool:
    """Return True when map output is missing or input digest changed."""
    sidecar = sidecar_path(output)
    if not output.exists() or not sidecar.exists():
        return True
    return sidecar.read_text().strip() != digest


def write_digest(output: Path, digest: str) -> None:
    """Write sidecar digest for a rendered map."""
    sidecar_path(output).write_text(digest + "\n")


def cleanup_orphan_maps(output_dir: Path, expected_stems: set[str]) -> int:
    """Delete orphaned PNGs and digest sidecars not backed by current inputs."""
    removed = 0
    if not output_dir.exists():
        return removed

    expected_names = {f"{stem}.png" for stem in expected_stems}
    expected_sidecars = {f"{stem}.sha256" for stem in expected_stems}

    for path in output_dir.iterdir():
        if not path.is_file():
            continue
        if path.suffix == ".png" and path.name not in expected_names:
            path.unlink()
            removed += 1
        elif path.suffix == ".sha256" and path.name not in expected_sidecars:
            path.unlink()
            removed += 1

    return removed


def render_map(
    m: Map,
    style: dict,
    locations: list[tuple[float, float]],
    output: Path,
    max_zoom: float = 13,
    zoom_out: float = 0.2,
) -> bool:
    """Render map with markers. Returns True if rendered, False if cached."""
    digest = map_input_digest(locations, max_zoom=max_zoom, zoom_out=zoom_out)
    if not needs_render(output, digest):
        return False

    geom = from_latlng(locations)

    # Calculate bounds from locations (lat, lng order)
    lats = [lat for lat, _lng in locations]
    lngs = [lng for _lat, lng in locations]
    bounds = (min(lngs), min(lats), max(lngs), max(lats))

    m.load_style(style)
    m.set_geojson("markers", geom)
    center, zoom = m.fit_bounds(bounds, max_zoom=max_zoom)
    png = m.render(center=center, zoom=zoom - zoom_out)
    output.write_bytes(png)
    write_digest(output, digest)
    return True


# --- Export functions ---


def _build_export_rows(companies: list[dict], taxonomies: dict) -> list[dict]:
    """Build export rows with display names for taxonomies (for CSV/XLSX readability)."""

    def keys_to_names(keys: list | None, tax: str) -> str:
        mapping = taxonomies.get(tax, {})
        return "; ".join(mapping.get(k, k) for k in (keys or []))

    rows = []
    for c in companies:
        slug = c.get("slug", "")

        # Compute logo_url
        logo_path = STATIC_DIR / "logos" / f"{slug}.png"
        logo_url = f"/logos/{slug}.png" if logo_path.exists() else ""

        # Compute overview_short
        overview = c.get("overview") or ""
        overview_short = overview[:150] + ("..." if len(overview) > 150 else "")

        # Compute search text
        search_text = (c.get("name", "") + " " + overview).lower().strip()

        rows.append(
            {
                "name": c.get("name", ""),
                "date": c.get("date", ""),
                "overview": overview,
                "overview_short": overview_short,
                "website": c.get("website", ""),
                "contact_name": c.get("contact_name", ""),
                "contact_title": c.get("contact_title", ""),
                "address": c.get("address", ""),
                "phone": c.get("phone", ""),
                "email": c.get("email", ""),
                "latitude": c.get("latitude"),
                "longitude": c.get("longitude"),
                "slug": slug,
                "logo_url": logo_url,
                "search": search_text,
                "company_types": keys_to_names(c.get("company_types"), "company_types"),
                "stakeholders": keys_to_names(c.get("stakeholders"), "stakeholders"),
                "capability_streams": keys_to_names(
                    c.get("capability_streams"), "capability_streams"
                ),
                "capability_domains": keys_to_names(
                    c.get("capability_domains"), "capability_domains"
                ),
                "industrial_capabilities": keys_to_names(
                    c.get("industrial_capabilities"), "industrial_capabilities"
                ),
                "regions": keys_to_names(c.get("regions"), "regions"),
                "ownerships": keys_to_names(c.get("ownerships"), "ownerships"),
                "content_html": md.render(c.get("_content", "").strip()),
            }
        )

    return rows


def sanitize_spreadsheet_cell(value: Any) -> Any:
    """Escape spreadsheet formula prefixes in text cells."""
    if not isinstance(value, str) or not value:
        return value
    if value[0] in ("=", "+", "-", "@", "	") or value[0] == "\0":
        return "'" + value
    return value


def sanitize_spreadsheet_df(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with spreadsheet-dangerous string cells escaped."""
    return df.map(sanitize_spreadsheet_cell)


def export_csv(df: pd.DataFrame, output: Path) -> None:
    """Export DataFrame to CSV."""
    sanitize_spreadsheet_df(df).to_csv(output, index=False)


def export_xlsx(df: pd.DataFrame, output: Path) -> None:
    """Export DataFrame to XLSX with formatting."""

    safe_df = sanitize_spreadsheet_df(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        safe_df.to_excel(writer, index=False, sheet_name="Companies")
        ws = writer.sheets["Companies"]

        # Hyperlinks for website column
        website_col = list(safe_df.columns).index("website") + 1
        for row_idx, url in enumerate(df["website"], start=2):
            if url and isinstance(url, str) and url.startswith("http"):
                cell = ws.cell(row=row_idx, column=website_col)
                cell.hyperlink = url
                cell.style = "Hyperlink"

        # Auto-width columns
        for idx, col in enumerate(safe_df.columns):
            col_data = safe_df[col].fillna("").astype(str)
            width = min(
                max(col_data.map(len).max(), len(col)) + 2,
                80 if col == "markdown" else 50,
            )
            ws.column_dimensions[get_column_letter(idx + 1)].width = width

        ws.freeze_panes = "A2"


def export_json(companies: list[dict], output: Path) -> None:
    """Export full JSON - all frontmatter fields plus computed fields."""
    data = []
    for c in companies:
        entry = {k: v for k, v in c.items() if not k.startswith("_")}

        # Add search text (lowercase name + overview)
        search_text = (
            entry.get("name", "") + " " + (entry.get("overview") or "")
        ).lower()
        entry["search"] = search_text.strip()

        # Add logo URL
        slug = entry.get("slug", "")
        logo_path = STATIC_DIR / "logos" / f"{slug}.png"
        entry["logo_url"] = f"/logos/{slug}.png" if logo_path.exists() else ""

        # Add truncated overview (150 chars)
        overview = entry.get("overview") or ""
        entry["overview_short"] = overview[:150] + (
            "..." if len(overview) > 150 else ""
        )

        # Add HTML content if present
        if c.get("_content"):
            entry["content_html"] = md.render(c["_content"].strip())

        data.append(entry)
    data.sort(key=lambda x: (x.get("name") or "").lower())
    output.write_text(json.dumps(data, indent=2))


def export_map_json(companies: list[dict], output: Path) -> None:
    """Export slim JSON for the interactive map only."""
    data = []
    for c in companies:
        slug = c.get("slug", "")
        logo_path = STATIC_DIR / "logos" / f"{slug}.png"
        overview = c.get("overview") or ""
        search_text = (c.get("name", "") + " " + overview).lower().strip()
        data.append(
            {
                "slug": slug,
                "name": c.get("name", ""),
                "overview": overview,
                "overview_short": overview[:150] + (
                    "..." if len(overview) > 150 else ""
                ),
                "logo_url": f"/logos/{slug}.png" if logo_path.exists() else "",
                "search": search_text,
                "latitude": c.get("latitude"),
                "longitude": c.get("longitude"),
                "stakeholders": c.get("stakeholders", []),
                "capability_streams": c.get("capability_streams", []),
                "capability_domains": c.get("capability_domains", []),
                "industrial_capabilities": c.get("industrial_capabilities", []),
                "regions": c.get("regions", []),
                "ownerships": c.get("ownerships", []),
                "company_types": c.get("company_types", []),
            }
        )
    data.sort(key=lambda x: (x.get("name") or "").lower())
    output.write_text(json.dumps(data, indent=2))


# --- Main ---


def main():
    companies = []
    company_locations: list[tuple[str, float, float]] = []
    term_locations: dict[str, dict[str, list[tuple[float, float]]]] = {
        tax: {} for tax in TAXONOMIES
    }

    # Load all companies
    for path in sorted(COMPANY_DIR.glob("*.md")):
        if path.name.startswith("_"):
            continue

        post = frontmatter.load(str(path))
        slug = path.stem
        fm = post.metadata

        # Store for export
        company = dict(fm)
        company["slug"] = slug
        company["_content"] = post.content
        companies.append(company)

        # Map locations
        lat, lng = fm.get("latitude"), fm.get("longitude")
        if lat and lng:
            company_locations.append((slug, float(lat), float(lng)))
            for taxonomy in TAXONOMIES:
                for term_key in fm.get(taxonomy, []):
                    term_locations[taxonomy].setdefault(term_key, []).append(
                        (float(lat), float(lng))
                    )

    print(f"Processed {len(companies)} companies")

    # Export files
    taxonomies = load_taxonomies(raw=True)
    STATIC_DIR.mkdir(parents=True, exist_ok=True)
    rows = _build_export_rows(companies, taxonomies)
    df = pd.DataFrame(rows).sort_values("name", ignore_index=True)
    export_csv(df, STATIC_DIR / "companies.csv")
    export_xlsx(df, STATIC_DIR / "companies.xlsx")
    export_json(companies, STATIC_DIR / "companies.json")
    export_map_json(companies, STATIC_DIR / "companies-map.json")
    print("Exported to CSV, XLSX, JSON, map JSON")

    # Generate maps
    if not company_locations:
        print("No coordinates, skipping maps")
        return

    print("Setting up map renderer...")
    m, style = setup_map()
    MAPS_DIR.mkdir(parents=True, exist_ok=True)
    TERM_MAPS_DIR.mkdir(parents=True, exist_ok=True)

    company_map_stems = {slug for slug, _lat, _lng in company_locations}
    removed_company = cleanup_orphan_maps(MAPS_DIR, company_map_stems)

    term_maps = [
        (f"{tax}-{key}", locs)
        for tax, terms in term_locations.items()
        for key, locs in terms.items()
        if locs
    ]
    term_map_stems = {key for key, _locs in term_maps}
    removed_term = cleanup_orphan_maps(TERM_MAPS_DIR, term_map_stems)

    with m:
        # Company maps
        count = 0
        with make_progress() as progress:
            task = progress.add_task("Company maps...", total=len(company_locations))
            for slug, lat, lng in company_locations:
                if render_map(m, style, [(lat, lng)], MAPS_DIR / f"{slug}.png"):
                    count += 1
                progress.update(task, advance=1)
        print(
            f"Generated {count} company maps ({len(company_locations) - count} cached, {removed_company} orphan files removed)"
        )

        # Term maps
        count = 0
        with make_progress() as progress:
            task = progress.add_task("Term maps...", total=len(term_maps))
            for key, locs in term_maps:
                if render_map(m, style, locs, TERM_MAPS_DIR / f"{key}.png"):
                    count += 1
                progress.update(task, advance=1)
        print(
            f"Generated {count} term maps ({len(term_maps) - count} cached, {removed_term} orphan files removed)"
        )


if __name__ == "__main__":
    main()
